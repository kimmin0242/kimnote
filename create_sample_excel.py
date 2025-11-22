#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
샘플 Excel 파일 생성 - 테스트용
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_sample_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "부품데이터"
    
    # Header
    headers = [
        '제조사', '대분류', '소분류', '모델명', '상세 트림/세대', '연식', '동력원 유형', '세부 엔진/동력계',
        '엔진오일(대)', '용량(L)', '엔진오일(소)', '용량(L)', '오일량 대+소', 
        '미션오일', '용량(L)', '브레이크오일', '용량(L)', 
        '에어필터', '오일필터', '에어컨 필터 실내', '에어컨 필터 외기', '배터리'
    ]
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Sample data - Genesis Sedans
    sample_data = [
        ['현대', '세단', '제네시스', 'G80', 'RG3', '2021-2024', '가솔린', '2.5 T-GDi', 
         '05100-00451', '6.5', '/', '/', '6.5L', '/', '/', '/', '/', 
         '28113-3N500', '26300-35530', '97133-D3000', '97133-D3100', '37110-M6000'],
        
        ['현대', '세단', '제네시스', 'G70', 'IK', '2022-2024', '가솔린', '2.0 터보', 
         '05100-00451', '5.0', '05100-00450', '0.5', '5.5L', 'HK SP-IV', '1.8', '/', '/', 
         '28113-2P100', '26300-35503', '[고급형/활성탄] 97133-D3000, [일반형] 97133-C1000', '97133-D3200', '37110-M0000'],
        
        ['현대', 'SUV', '제네시스', 'GV70', 'JK1', '2021-2024', '디젤', '2.2 디젤', 
         '05100-00470', '7.0', '/', '/', '7.0L', '/', '/', 'DOT 4', '0.5', 
         '28113-3N500', '26320-2F300', '97133-D3000', '97133-D3300', '37110-M6000'],
        
        ['현대', '승용차', '현대', '그랜저', 'GN7', '2023-2024', '가솔린', '3.5 GDi', 
         '05100-00451', '6.0', '/', '/', '6.0L', '/', '/', '/', '/', 
         '28113-3L100', '26300-35530', '97133-L1000', '97133-L1100', '37110-2S000'],
        
        ['현대', 'SUV', '현대', '팰리세이드', 'LX2', '2020-2024', '디젤', '2.2 디젤', 
         '05100-00470', '7.0', '05100-00471', '0.5', '7.5L', 'SP-IV-M', '2.0', 'DOT 4', '0.5', 
         '28113-S8100', '26320-2F100', '97133-S8100', '97133-S8200', '37110-3N500'],
        
        ['현대', '전기차', '현대', '아이오닉5', '/', '2023-2024', '전기', '58kWh 배터리', 
         '/', '/', '/', '/', '/', '/', '/', 'DOT 4', '0.3', 
         '28113-GI000', '/', '97133-GI100', '97133-GI200', '37110-IONIQ5'],
        
        ['현대', '수소차', '현대', '넥쏘', '/', '2023-2024', '수소', '수소연료전지', 
         '/', '/', '/', '/', '/', '/', '/', '/', '/', 
         '28113-N9000', '/', '97133-N9000', '97133-N9100', '37110-NEXO'],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18
    
    # Save
    filename = 'sample_parts_data.xlsx'
    wb.save(filename)
    print(f"✅ 샘플 파일 생성 완료: {filename}")
    print(f"   - 총 {len(sample_data)}개 차량 데이터")
    print(f"   - 제네시스 세단: G80, G70")
    print(f"   - 제네시스 SUV: GV70")
    print(f"   - 현대 승용차: 그랜저")
    print(f"   - 현대 SUV: 팰리세이드")
    print(f"   - 현대 친환경: 아이오닉5, 넥쏘")

if __name__ == '__main__':
    create_sample_excel()
