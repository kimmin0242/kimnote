#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 데이터 변환 도구 - 방법 2: 차종별 시트 분리
Converts horizontal vehicle parts data to vertical format with separate sheets by vehicle type

사용법:
python3 excel_converter_method2.py input.xlsx output.xlsx
"""

import sys
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# 차종별 시트 매핑
VEHICLE_TYPE_SHEETS = {
    '제너시스_세단': {
        'brands': ['제네시스', 'GENESIS', 'Genesis'],
        'categories': ['세단', '승용차'],
        'models': ['G90', 'G80', 'G70']
    },
    '제너시스_SUV': {
        'brands': ['제네시스', 'GENESIS', 'Genesis'],
        'categories': ['SUV'],
        'models': ['GV60', 'GV70', 'GV80', 'GV90']
    },
    '현대_승용차': {
        'brands': ['현대', 'HYUNDAI', 'Hyundai', '현대자동차'],
        'categories': ['세단', '승용차', '해치백'],
        'models': ['아반떼', '벨로스터', '쏘나타', '그랜저', '에쿠스', 'i30', 'i40']
    },
    '현대_SUV': {
        'brands': ['현대', 'HYUNDAI', 'Hyundai', '현대자동차'],
        'categories': ['SUV', 'RV'],
        'models': ['코나', '투싼', '싼타페', '팰리세이드', '베뉴', '넥쏘', '맥스크루즈', 'ix35', 'ix55']
    },
    '현대_친환경차': {
        'brands': ['현대', 'HYUNDAI', 'Hyundai', '현대자동차', '제네시스', 'GENESIS'],
        'categories': ['전기차', 'EV', '수소차', 'FCEV', '친환경'],
        'models': ['아이오닉5', '아이오닉6', 'IONIQ5', 'IONIQ6', '넥쏘', 'NEXO', 'GV60']
    }
}

# 부품 분류 매핑 (부품명 -> 대분류, 소분류)
PART_CATEGORY_MAPPING = {
    '엔진오일(대)': ('오일 및 액체류', '엔진오일(대)'),
    '엔진오일(소)': ('오일 및 액체류', '엔진오일(소)'),
    '미션오일': ('오일 및 액체류', '미션오일'),
    '브레이크오일': ('오일 및 액체류', '브레이크오일'),
    '냉각수/부동액': ('오일 및 액체류', '냉각수/부동액'),
    '부동액': ('오일 및 액체류', '냉각수/부동액'),
    '파워스티어링오일': ('오일 및 액체류', '파워스티어링오일'),
    '워셔액': ('오일 및 액체류', '워셔액'),
    '디퍼런셜오일': ('오일 및 액체류', '디퍼런셜오일'),
    
    '에어필터': ('필터류', '에어필터'),
    '오일필터': ('필터류', '오일필터'),
    '에어컨필터(실내)': ('필터류', '에어컨필터(실내)'),
    '에어컨 필터 실내': ('필터류', '에어컨필터(실내)'),
    '에어컨필터(외기)': ('필터류', '에어컨필터(외기)'),
    '에어컨 필터 외기': ('필터류', '에어컨필터(외기)'),
    '연료필터': ('필터류', '연료필터'),
    
    '브레이크 패드(앞축)': ('제동류', '브레이크 패드(앞축)'),
    '브레이크 패드(뒤축)': ('제동류', '브레이크 패드(뒤축)'),
    '브레이크 디스크(앞)': ('제동류', '브레이크 디스크(앞)'),
    '브레이크 디스크(뒤)': ('제동류', '브레이크 디스크(뒤)'),
    '타이어(앞)': ('제동류', '타이어(앞)'),
    '타이어(뒤)': ('제동류', '타이어(뒤)'),
    
    '배터리': ('전장 및 기타 부품류', '배터리'),
    '점화플러그': ('전장 및 기타 부품류', '점화플러그'),
    '점화코일': ('전장 및 기타 부품류', '점화코일'),
    '구동벨트 (V벨트)': ('전장 및 기타 부품류', '구동벨트 (V벨트)'),
    '타이밍벨트': ('전장 및 기타 부품류', '타이밍벨트'),
    '와이퍼 블레이드(좌)': ('전장 및 기타 부품류', '와이퍼 블레이드(좌)'),
    '와이퍼 블레이드(우)': ('전장 및 기타 부품류', '와이퍼 블레이드(우)'),
    '와이퍼 블레이드(뒤)': ('전장 및 기타 부품류', '와이퍼 블레이드(뒤)'),
}

def clean_value(value):
    """Clean cell value"""
    if value is None:
        return None
    
    str_value = str(value).strip()
    
    # Skip empty or "/" values
    if str_value in ['', '/', '-', 'N/A', 'n/a']:
        return None
    
    return str_value

def determine_vehicle_sheet(manufacturer, category, model, subcategory=''):
    """Determine which sheet this vehicle belongs to"""
    manufacturer = clean_value(manufacturer) or ''
    category = clean_value(category) or ''
    model = clean_value(model) or ''
    subcategory = clean_value(subcategory) or ''
    
    # Combine manufacturer and subcategory for brand detection
    brand_str = f"{manufacturer} {subcategory}".lower()
    
    for sheet_name, criteria in VEHICLE_TYPE_SHEETS.items():
        # Check manufacturer/brand (check both manufacturer and subcategory fields)
        brand_match = any(brand.lower() in brand_str for brand in criteria['brands'])
        
        # Check category
        category_match = any(cat in category for cat in criteria['categories'])
        
        # Check model
        model_match = any(m in model for m in criteria['models'])
        
        # For 친환경차, prioritize specific models or categories
        if sheet_name == '현대_친환경차':
            if model_match or any(cat in category for cat in ['전기차', 'EV', '수소차', 'FCEV', '친환경']):
                return sheet_name
        
        # For other sheets, require brand match + (category or model match)
        if brand_match and (category_match or model_match):
            # Exclude 친환경 models from regular sheets
            if model not in ['아이오닉5', '아이오닉6', 'IONIQ5', 'IONIQ6', '넥쏘', 'NEXO']:
                return sheet_name
    
    # Default fallback
    if '제네시스' in brand_str or 'genesis' in brand_str:
        if 'SUV' in category or 'GV' in model:
            return '제너시스_SUV'
        return '제너시스_세단'
    else:
        if 'SUV' in category or 'RV' in category:
            return '현대_SUV'
        return '현대_승용차'

def extract_part_info(part_number_str):
    """Extract part options from part number string like '[고급형/활성탄] 12345, [일반형] 67890'"""
    if not part_number_str:
        return []
    
    parts = []
    # Split by comma
    segments = str(part_number_str).split(',')
    
    for segment in segments:
        segment = segment.strip()
        if not segment or segment == '/':
            continue
        
        option = ''
        part_number = segment
        
        # Extract option in brackets [...]
        if '[' in segment and ']' in segment:
            start = segment.index('[')
            end = segment.index(']')
            option = segment[start+1:end].strip()
            part_number = segment[end+1:].strip()
        
        parts.append({
            'option': option,
            'part_number': part_number
        })
    
    return parts

def create_styled_header(ws):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Freeze header row
    ws.freeze_panes = 'A2'

def convert_horizontal_to_vertical(input_file, output_file):
    """Convert horizontal Excel format to vertical format with separate sheets"""
    
    print(f"📂 입력 파일 로딩: {input_file}")
    
    if not os.path.exists(input_file):
        print(f"❌ 오류: 입력 파일을 찾을 수 없습니다: {input_file}")
        return False
    
    try:
        wb_input = load_workbook(input_file, data_only=True)
        ws_input = wb_input.active
        
        print(f"✓ 시트 로딩 완료: {ws_input.title}")
        print(f"  총 행 수: {ws_input.max_row}")
        print(f"  총 열 수: {ws_input.max_column}")
        
    except Exception as e:
        print(f"❌ 파일 로딩 오류: {e}")
        return False
    
    # Create output workbook
    wb_output = Workbook()
    wb_output.remove(wb_output.active)  # Remove default sheet
    
    # Create sheets for each vehicle type
    sheets = {}
    for sheet_name in VEHICLE_TYPE_SHEETS.keys():
        ws = wb_output.create_sheet(title=sheet_name)
        
        # Header row
        headers = [
            '제조사', '브랜드', '카테고리', '모델명', '세대', '연식', '연료타입', '엔진타입',
            '부품분류', '부품타입', '부품번호', '용량(L)', '수량', '옵션', '비고'
        ]
        ws.append(headers)
        create_styled_header(ws)
        
        # Set column widths
        ws.column_dimensions['A'].width = 10  # 제조사
        ws.column_dimensions['B'].width = 12  # 브랜드
        ws.column_dimensions['C'].width = 10  # 카테고리
        ws.column_dimensions['D'].width = 12  # 모델명
        ws.column_dimensions['E'].width = 10  # 세대
        ws.column_dimensions['F'].width = 10  # 연식
        ws.column_dimensions['G'].width = 12  # 연료타입
        ws.column_dimensions['H'].width = 18  # 엔진타입
        ws.column_dimensions['I'].width = 15  # 부품분류
        ws.column_dimensions['J'].width = 18  # 부품타입
        ws.column_dimensions['K'].width = 15  # 부품번호
        ws.column_dimensions['L'].width = 10  # 용량
        ws.column_dimensions['M'].width = 8   # 수량
        ws.column_dimensions['N'].width = 15  # 옵션
        ws.column_dimensions['O'].width = 20  # 비고
        
        sheets[sheet_name] = ws
    
    # Read header from input
    input_headers = []
    for cell in ws_input[1]:
        input_headers.append(clean_value(cell.value) or '')
    
    print(f"\n📋 입력 파일 헤더:")
    for i, h in enumerate(input_headers):
        print(f"  [{i}] {h}")
    
    # Map column indices
    col_map = {}
    for i, header in enumerate(input_headers):
        col_map[header] = i
    
    # Statistics
    stats = {sheet_name: 0 for sheet_name in VEHICLE_TYPE_SHEETS.keys()}
    total_rows_processed = 0
    skipped_rows = 0
    
    # Process each row
    print(f"\n🔄 데이터 변환 시작...\n")
    
    for row_idx in range(2, ws_input.max_row + 1):
        row = ws_input[row_idx]
        row_data = [cell.value for cell in row]
        
        # Extract basic vehicle info
        manufacturer = clean_value(row_data[col_map.get('제조사', 0)])
        category = clean_value(row_data[col_map.get('대분류', 1)])
        subcategory = clean_value(row_data[col_map.get('소분류', 2)])
        model_name = clean_value(row_data[col_map.get('모델명', 3)])
        generation = clean_value(row_data[col_map.get('상세 트림/세대', 4)])
        year = clean_value(row_data[col_map.get('연식', 5)])
        fuel_type = clean_value(row_data[col_map.get('동력원 유형', 6)])
        engine_type = clean_value(row_data[col_map.get('세부 엔진/동력계', 7)])
        
        # Skip if no model name
        if not model_name:
            skipped_rows += 1
            continue
        
        # Determine target sheet
        target_sheet_name = determine_vehicle_sheet(manufacturer, category, model_name, subcategory)
        target_ws = sheets[target_sheet_name]
        
        # Process each part type
        parts_found = 0
        
        # Find all part columns (starting from column 8)
        for col_idx in range(8, len(input_headers)):
            part_type = input_headers[col_idx]
            
            # Skip capacity and notes columns (they are processed with their parts)
            if not part_type or '용량' in part_type or part_type in ['비고', '오일량', '오일량 대+소']:
                continue
            
            part_number_raw = clean_value(row_data[col_idx])
            
            if not part_number_raw:
                continue
            
            # Get category for this part type
            part_category = PART_CATEGORY_MAPPING.get(part_type, ('기타', part_type))
            
            # Extract part info (handles multiple options)
            part_infos = extract_part_info(part_number_raw)
            
            if not part_infos:
                # Simple part number without options
                part_infos = [{'option': '', 'part_number': part_number_raw}]
            
            # Get capacity and notes from next columns
            capacity = ''
            notes = ''
            
            # Try to find capacity column after current part
            # Check next 2 columns for capacity data
            for offset in [1, 2]:
                next_idx = col_idx + offset
                if next_idx < len(input_headers):
                    next_header = input_headers[next_idx]
                    if next_header and '용량' in next_header:
                        capacity_value = clean_value(row_data[next_idx]) if next_idx < len(row_data) else ''
                        if capacity_value:
                            capacity = capacity_value
                        break
            
            # Add row for each part option
            for part_info in part_infos:
                target_ws.append([
                    manufacturer or '현대',
                    subcategory or category,  # 브랜드
                    category,  # 카테고리
                    model_name,
                    generation or '',
                    year or '',
                    fuel_type or '',
                    engine_type or '',
                    part_category[0],  # 부품분류
                    part_category[1],  # 부품타입
                    part_info['part_number'],
                    capacity,
                    '',  # 수량
                    part_info['option'],
                    notes
                ])
                
                parts_found += 1
        
        if parts_found > 0:
            stats[target_sheet_name] += parts_found
            total_rows_processed += 1
            
            if total_rows_processed % 10 == 0:
                print(f"  처리 중... {total_rows_processed}개 차량 처리됨")
    
    # Save output file
    print(f"\n💾 출력 파일 저장 중: {output_file}")
    
    try:
        wb_output.save(output_file)
        print(f"✅ 변환 완료!\n")
        
        # Print statistics
        print("📊 변환 통계:")
        print(f"  총 처리 차량: {total_rows_processed}")
        print(f"  건너뛴 행: {skipped_rows}")
        print(f"\n  시트별 부품 수:")
        for sheet_name, count in stats.items():
            print(f"    {sheet_name}: {count}개")
        
        total_parts = sum(stats.values())
        print(f"\n  총 부품 레코드: {total_parts}개")
        
        return True
        
    except Exception as e:
        print(f"❌ 저장 오류: {e}")
        return False

def main():
    if len(sys.argv) < 2:
        print("""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Excel 데이터 변환 도구 - 방법 2: 차종별 시트 분리
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

사용법:
  python3 excel_converter_method2.py <입력파일> [출력파일]

예시:
  python3 excel_converter_method2.py 부품데이터.xlsx
  python3 excel_converter_method2.py 부품데이터.xlsx 변환결과.xlsx

출력 시트:
  • 제너시스_세단 (G90, G80, G70)
  • 제너시스_SUV (GV60, GV70, GV80)
  • 현대_승용차 (아반떼, 쏘나타, 그랜저 등)
  • 현대_SUV (코나, 투싼, 싼타페, 팰리세이드 등)
  • 현대_친환경차 (아이오닉, 넥쏘 등)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """)
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    # Generate output filename
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        base_name = os.path.splitext(input_file)[0]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"{base_name}_차종별시트_{timestamp}.xlsx"
    
    print(f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Excel 데이터 변환 시작
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
입력: {input_file}
출력: {output_file}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    """)
    
    success = convert_horizontal_to_vertical(input_file, output_file)
    
    if success:
        print(f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  ✅ 변환 성공!
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
생성된 파일: {output_file}

다음 단계:
1. 생성된 파일을 열어서 데이터 확인
2. 각 시트별로 데이터가 올바르게 분류되었는지 검토
3. 필요시 수동으로 조정
4. admin_import_data.php로 데이터베이스 임포트

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """)
        sys.exit(0)
    else:
        print("\n❌ 변환 실패")
        sys.exit(1)

if __name__ == '__main__':
    main()
